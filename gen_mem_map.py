#!/eda/python3.7/bin/python3
import pandas as pd
import os
import argparse
from openpyxl import load_workbook
class gen_mem_map:
	def __init__(self):
		self.c_header_content =[]
		self.c_header_mem_size =[]
		self.v_define_content =[]

		self.block_column_name = 'G'
		self.memory_column_num = 11
		self.size_column_num  = 9

	def arg_h(self):
		parser = argparse.ArgumentParser('description= memory map generation script')
		parser.add_argument('-source',help='memory map xlsx file path and name',action='store',dest='source_file',default='/proj/HPM1200/spec/mem_map/mem_map_latest.xlsx')
		args = parser.parse_args()
		self.source_file = args.source_file
	
	def process_c_header(self, addr_column_num, irq_sheet, dma_sheet):
		wb = load_workbook(self.source_file)
		ws = wb['MEM_MAP']
		row_index = 0
		for block in ws[self.block_column_name]:
			row_index += 1
			if row_index == 1:
				pass
			elif block.value.upper() == 'RESERVED':
				continue
			else:
				block_name = block.value.upper()
				addr = ws.cell(row=row_index, column=addr_column_num).value
				if addr == None:
					continue
				self.c_header_content.append('%-8s%-30s%-12s\n'%('#define MAP_BASE_', block_name, addr))
				if(ws.cell(row=row_index, column=self.memory_column_num).value != None):
					mem_size = ws.cell(row=row_index, column=self.size_column_num).value.upper()
					if 'KB' in mem_size:
						mem_size = int(mem_size.replace('KB',''))*1024
					elif 'MB' in mem_size:
						mem_size = int(mem_size.replace('MB',''))*1024*1024
					mem_size = '0x'+str('%x'%mem_size)
					self.c_header_mem_size.append('%-8s%-30s%-12s\n'%('#define SIZE_', block_name, mem_size))
		self.c_header_content.append('\n\n')
		self.c_header_content.append('//Used by files from SDK: hpm_plic_drv.h, hpm_interrupt.h \n')
		self.c_header_content.append('#define HPM_PLIC_BASE    MAP_BASE_PLIC \n')
		self.c_header_content.append('#define HPM_PLICSW_BASE  MAP_BASE_PLIC_SW \n')
		self.c_header_content.append("\n\n")
		ws = wb[irq_sheet]
		row_index = 0
		for irq in ws['C']:
			row_index += 1
			if row_index == 1:
				pass
			else:
				if irq.value != None:
					irq_name = irq.value.upper()
					irq_name = irq_name.replace('[','_')
					irq_name = irq_name.replace(']','')
					irq_index = ws.cell(row=row_index, column=2).value
					self.c_header_content.append('%-8s%-20s%-8s\n'%('#define IRQ_INDEX_', irq_name, irq_index))
		self.c_header_content.append("\n\n")
		ws = wb[dma_sheet]
		row_index = 0
		for dma in ws['C']:
			row_index += 1
			if row_index == 1:
				pass
			else:
				dma_name = dma.value.upper()
				dma_name = dma_name.replace('[','_')
				dma_name = dma_name.replace(']','')
				dma_index = ws.cell(row=row_index, column=2).value
				self.c_header_content.append('%-8s%-20s%-8s\n'%('#define DMA_INDEX_', dma_name, dma_index))

	def process_v_define(self, addr_column_num, platform_num):
		wb = load_workbook(self.source_file)
		ws = wb['MEM_MAP']
		row_index = 0
		for block in ws[self.block_column_name]:
			row_index += 1
			if row_index == 1:
				pass
			elif block.value.upper() == 'RESERVED':
				continue
			else:
				if(ws.cell(row=row_index, column=self.memory_column_num).value != None):
					block_name = block.value.upper()
					for i_column_shift in range(platform_num):
						addr = ws.cell(row=row_index, column=(addr_column_num+i_column_shift)).value
						if addr != None:
							break
					addr = str(addr).replace('0x',"32'h")
					self.v_define_content.append('%-8s%-30s%-12s\n'%('`define MAP_BASE_', block_name, addr))
					mem_size = ws.cell(row=row_index, column=self.size_column_num).value.upper()
					if 'KB' in mem_size:
						mem_size = int(mem_size.replace('KB',''))*1024
					elif 'MB' in mem_size:
						mem_size = int(mem_size.replace('MB',''))*1024*1024
					mem_size = "32'h"+str('%x'%mem_size)
					self.v_define_content.append('%-8s%-30s%-12s\n'%('`define SIZE_', block_name, mem_size))
		self.v_define_content.append("\n\n")
		self.v_define_content.append("\n\n")
		self.v_define_content.append("`define MAP_BASE_FUSE_MEM  (`MAP_BASE_FUSE + 32'h400)\n`define SIZE_FUSE_MEM 32'h0000_0200 \n")
		self.v_define_content.append("\n\n")

	def run(self):
		#SOC memory map C-header
		self.process_c_header(8, 'IRQ', 'DMA')
		f = open('../verification/common_c/system/include/soc/sys_mem_map.h','w')
		f.write('//This file is generated automaticly, please do not edit it manually\n')
		f.write("#ifndef __SYS_MEM_MAP__H__\n")
		f.write("#define __SYS_MEM_MAP__H__\n\n\n")
		for i in self.c_header_content:
			f.write(i)
		f.write("\n\n")
		for i in self.c_header_mem_size:
			f.write(i)
		f.write("\n\n")
		f.write('#define MAP_BASE_CVC   (MAP_BASE_MISC)\n')
		f.write('\n#endif //__SYS_MEM_MAP__H__')
		f.close()
		#V-define
		self.process_v_define(8, 1)
		f = open('../verification/testbench/define/tb_mem_map.v','w')
		f.write('//This file is generated automaticly, please do not edit it manually\n')
		for i in self.v_define_content:
			f.write(i)
		f.write("\n\n")
		f.close()

if __name__ == '__main__':
	a=gen_mem_map()
	a.arg_h()
	a.run()



