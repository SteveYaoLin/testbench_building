#!/eda/python3.7/bin/python3
import argparse
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import pandas as pd
from fnmatch import fnmatch, fnmatchcase
import re

class gen_tb_io:

	#def __init__(self):
		#self.df=pd.DataFrame()
		#self.df_tmp = pd.DataFrame(columns=['name','0','1','2','3','4','5','6','current','flag','sub_flag','val'])
		#self.df_out = pd.DataFrame()
		#self.module_list = []
	def arg_h(self):
		parser = argparse.ArgumentParser('description= pinmux header generation script')
		parser.add_argument('-source',help='pinmux xlsx file path and name',action='store',dest='source_file',default='/proj/HPM1200/spec/pin_mux/pin_mux_latest.xlsx')
		args = parser.parse_args()
		self.source_file = args.source_file
	
	def instance_rename(self, instance_name):
		result = instance_name.replace('DIS','LCDC')
		result = result.replace('ETH','ENET')
		result = result.replace('SDC','SDXC')
		result = result.replace('URT','UART')
		result = result.replace('MIPI0','MIPI_CSI0')
		result = result.replace('MIPI1','MIPI_CSI1')
		result = result.replace('MIPI2','MIPI_DSI0')
		result = result.replace('MIPI3','MIPI_DSI1')
		result = result.replace('LVDS0','LVDS_RX0')
		result = result.replace('LVDS1','LVDS_RX1')
		result = result.replace('LVDS2','LVDS_TX0')
		result = result.replace('LVDS3','LVDS_TX1')
		return result

	def get_opt_list(self, option):
		result = []
		#ASCII code: A-65 B-66 P-80
		for each_char in option:
			assii_code = ord(each_char)
			if assii_code<65 or assii_code>80:
				print('Error: %s is not a valid option letter',each_char)
			else:
				result.append('opt'+str(assii_code-65))
		return result


	def run(self):
		wb = load_workbook(self.source_file, data_only=True)
		#instance_list = []
		#pinmux_table = pd.DataFrame(columns=['module','instance','pad_func','alt','opt0','opt1','opt2','opt3','opt4','opt5','opt6','opt7'])
		pinmux_table_column_list = ['module','instance','pad_func','alt','domain']
		opt_list = ['opt0','opt1','opt2','opt3','opt4','opt5','opt6','opt7','opt8','opt9','opt10','opt11','opt12','opt13','opt14','opt15'] 
		for i_column in opt_list:
			pinmux_table_column_list.append(i_column)
		pinmux_table = pd.DataFrame(columns=pinmux_table_column_list)
		#print(pinmux_table_column_list)
		#pinmux_soc
		ws = wb['PINMUX']
		row_index = 0
		for ballname in ws['C']: #column I is "BALLNAME"
			row_index += 1
			#name = ballname.value
			if row_index == 1:
				if ballname.value != 'BALLNAME':
					print('Error column')
			elif ballname.value == None:
				break
			elif fnmatch(ballname.value, '[PA]*'):
				column_base = column_index_from_string('H') #ALT0(GPIO)
				for column_shift in range(1,32):
					func = ws.cell(row=row_index, column=(column_base+column_shift)).value
					if func != None:
						instance = func.split('.')[0]
						instance = self.instance_rename(instance)
						#if instance in instance_list:
						#	pass
						#else:
						#	instance_list.append(instance)
						#remove end-numbers from instance
						module = re.sub(r'[0-9]+$', '', instance)
						#pinmux_table.at[pinmux_table_row_index,'module'] = re.sub(r'[0-9]+$', '', instance)
						#pinmux_table.at[pinmux_table_row_index,'instance'] = instance
						pad_func = func.split('.')[2]
						pad_func = pad_func.replace('[','')
						pad_func = pad_func.replace(']','')
						#pinmux_table.at[pinmux_table_row_index,'pad_func'] = pad_func
						if len(pinmux_table[(pinmux_table.module==module)&(pinmux_table.instance==instance)&(pinmux_table.pad_func==pad_func)].index.tolist()) == 0:
							pinmux_table.loc[len(pinmux_table)] = {'module':module,'instance':instance,'pad_func':pad_func,'alt':column_shift,'domain':'soc'}
						option = func.split('.')[1]
						pad_opt_list = self.get_opt_list(option)
						pinmux_table_row_index = pinmux_table[(pinmux_table.module==module)&(pinmux_table.instance==instance)&(pinmux_table.pad_func==pad_func)].index.tolist()[0]
						for i_opt_value in pad_opt_list:
						 	pinmux_table.at[pinmux_table_row_index, i_opt_value] = ballname.value
				for column_shift in range(0,8):
					func = ws.cell(row=row_index, column=(column_index_from_string('AW')+column_shift)).value #column BF is ADC0 analog function
					if func != None:
						instance = func.split('.')[0]
						instance = self.instance_rename(instance)
						#if instance in instance_list:
						#	pass
						#else:
						#	instance_list.append(instance)
						#remove end-numbers from instance
						module = re.sub(r'[0-9]+$', '', instance)
						#pinmux_table.at[pinmux_table_row_index,'module'] = re.sub(r'[0-9]+$', '', instance)
						#pinmux_table.at[pinmux_table_row_index,'instance'] = instance
						pad_func = func.split('.')[2]
						pad_func = pad_func.replace('[','')
						pad_func = pad_func.replace(']','')
						#pinmux_table.at[pinmux_table_row_index,'pad_func'] = pad_func
						if len(pinmux_table[(pinmux_table.module==module)&(pinmux_table.instance==instance)&(pinmux_table.pad_func==pad_func)].index.tolist()) == 0:
							pinmux_table.loc[len(pinmux_table)] = {'module':module,'instance':instance,'pad_func':pad_func,'alt':32,'domain':'soc'}
						option = func.split('.')[1]
						pad_opt_list = self.get_opt_list(option)
						pinmux_table_row_index = pinmux_table[(pinmux_table.module==module)&(pinmux_table.instance==instance)&(pinmux_table.pad_func==pad_func)].index.tolist()[0]
						for i_opt_value in pad_opt_list:
						 	pinmux_table.at[pinmux_table_row_index, i_opt_value] = ballname.value
		#pinmux_pmic
		ws = wb['PINMUX_PMIC']
		row_index = 0
		for ballname in ws['I']: #column I is "BALLNAME"
			row_index += 1
			#name = ballname.value
			if row_index == 1:
				if ballname.value != 'BALLNAME':
					print('Error column')
			elif ballname.value == None:
				break
			elif fnmatch(ballname.value, 'PY*'):
				column_base = column_index_from_string('N') #column N is ALT0(GPIO)
				for column_shift in range(1,3):
					func = ws.cell(row=row_index, column=(column_base+column_shift)).value
					if func != None:
						instance = func.split('.')[0]
						instance = self.instance_rename(instance)
						#remove end-numbers from instance
						module = re.sub(r'[0-9]+$', '', instance)
						#remove "P" from instance
						module = re.sub(r'^P', '', instance)
						pad_func = func.split('.')[2]
						pad_func = pad_func.replace('[','')
						pad_func = pad_func.replace(']','')
						if len(pinmux_table[(pinmux_table.module==module)&(pinmux_table.instance==instance)&(pinmux_table.pad_func==pad_func)].index.tolist()) == 0:
							pinmux_table.loc[len(pinmux_table)] = {'module':module,'instance':instance,'pad_func':pad_func,'alt':column_shift,'domain':'pmic'}
						option = func.split('.')[1]
						pad_opt_list = self.get_opt_list(option)
						pinmux_table_row_index = pinmux_table[(pinmux_table.module==module)&(pinmux_table.instance==instance)&(pinmux_table.pad_func==pad_func)].index.tolist()[0]
						for i_opt_value in pad_opt_list:
						 	pinmux_table.at[pinmux_table_row_index, i_opt_value] = ballname.value

		#pinmux_batt
		ws = wb['PINMUX_BATT']
		row_index = 0
		for ballname in ws['I']: #column I is "BALLNAME"
			row_index += 1
			#name = ballname.value
			if row_index == 1:
				if ballname.value != 'BALLNAME':
					print('Error column')
			elif ballname.value == None:
				break
			elif fnmatch(ballname.value, 'PZ*'):
				column_base = column_index_from_string('N') #column N is ALT0(GPIO)
				for column_shift in range(1,3):
					func = ws.cell(row=row_index, column=(column_base+column_shift)).value
					#print (column_shift)
					if func != None:
						instance = func.split('.')[0]
						instance = self.instance_rename(instance)
						#remove end-numbers from instance
						module = re.sub(r'[0-9]+$', '', instance)
						pad_func = func.split('.')[2]
						pad_func = pad_func.replace('[','')
						pad_func = pad_func.replace(']','')
						if len(pinmux_table[(pinmux_table.module==module)&(pinmux_table.instance==instance)&(pinmux_table.pad_func==pad_func)].index.tolist()) == 0:
							pinmux_table.loc[len(pinmux_table)] = {'module':module,'instance':instance,'pad_func':pad_func,'alt':column_shift,'domain':'batt'}
						option = func.split('.')[1]
						pad_opt_list = self.get_opt_list(option)
						pinmux_table_row_index = pinmux_table[(pinmux_table.module==module)&(pinmux_table.instance==instance)&(pinmux_table.pad_func==pad_func)].index.tolist()[0]
						for i_opt_value in pad_opt_list:
						 	pinmux_table.at[pinmux_table_row_index, i_opt_value] = ballname.value

		pinmux_table = pinmux_table.sort_values(by=['module','instance','pad_func']) # sort by "module" then "instance" column then "pad_func"
		pinmux_table = pinmux_table.reset_index(drop=True) # reset the index as 0,1,2,3...
		#add a dummy line to make sure the last module will be processed by later program
		pinmux_table.loc[len(pinmux_table)] = {'module':'DUMMY','instance':'DUMMY','pad_func':'DUMMY'}
		pd.set_option('display.max_rows', None)
		pd.set_option('display.max_columns', None)
		pd.set_option('display.width', None)
		#print (pinmux_table)


		temp_table = pinmux_table.copy(deep=True)
		#clear the temp_table
		temp_table.drop(temp_table.index,inplace=True)
		current_module = ''
		current_instance = ''
		#c_write_content = []
		#pinmux_table_row_index = 0
		temp_table_row_index = 0
		#for i_module in pinmux_table.loc[:, 'module'] :
		c_include_file = open('../verification/common_c/system/include/sys_io_config.h','w')
		c_include_file.write('#ifndef __SYS_IO_CONFIG__H__\n')
		c_include_file.write('#define __SYS_IO_CONFIG__H__\n\n')
		c_include_file.write('#include \"io_config/sys_inst_map.h\"\n\n')
		v_include_file = open('../verification/testbench/wire/tb_connect_autogen_wrapper.sv','w')
		pinmux_summary_file = open('../verification/spec/pinmux_summary','w')
		for pinmux_table_row_index, pinmux_table_row in pinmux_table.iterrows() :
			if current_module == '' :
				current_module = pinmux_table_row['module']
			if pinmux_table_row['module'] == current_module:
				#temp_array = pinmux_table.loc[pinmux_table_row_index]
				temp_table = temp_table.append(pinmux_table_row, ignore_index=True)
			else :
				#temp_table = temp_table.sort_values(by=['instance','opt0','opt1','opt2','opt3','opt4','opt5','opt6','opt7'])
				#temp_table = temp_table.reset_index(drop=True) # reset the index as 0,1,2,3...
				print(temp_table, file=pinmux_summary_file)
				c_header_file_name = '../verification/common_c/system/include/io_config/io_config_' + current_module.lower() + '.h'
				c_header_file_name_short = 'io_config_' + current_module.lower() + '.h'
				v_wire_file_name = '../verification/testbench/wire/tb_connect_' + current_module.lower() + '.sv'
				#print (c_header_file_name)
				temp_table_row_index = 0
				c_write_content = []
				v_write_content = []
				#c_header_file = open(c_header_file_name,'w')
				v_wire_file = open(v_wire_file_name,'w')
				wire_list = []
				for index, temp_table_row in temp_table.iterrows() :
					temp_wire = str('tb_'+temp_table_row['instance'].lower()+'_'+temp_table_row['pad_func'].lower())
					if temp_wire in wire_list:
						pass
					else:
						v_write_content.append('wire '+temp_wire+';\n')
						wire_list.append(temp_wire)
				for i_opt in opt_list:
					current_instance = ''
					for temp_table_row_index, temp_table_row in temp_table.iterrows() :
						if str(temp_table_row[i_opt]) != 'nan':
							if current_instance != temp_table_row['instance'].lower() :
								instance_name = temp_table_row['instance']
								v_write_content.append('\n') 
								v_write_content.append('logic active_'+instance_name.lower()+'_'+i_opt+' = 1\'b0;\n')
								v_write_content.append('always begin\n')
								v_write_content.append('    `HIER_TB_CVC_SOC.cvc_trigger_get(`TRIGGER_TB_CONNECT_'+instance_name+'_'+i_opt.upper()+');\n')
								v_write_content.append('    active_'+instance_name.lower()+'_'+i_opt+' = 1\'b1;\n')
								v_write_content.append('end\n')
								v_write_content.append('always begin\n')
								v_write_content.append('    `HIER_TB_CVC_SOC.cvc_trigger_get(`TRIGGER_TB_DISCONNECT_'+instance_name+'_'+i_opt.upper()+');\n')
								v_write_content.append('    active_'+instance_name.lower()+'_'+i_opt+' = 1\'b0;\n')
								v_write_content.append('end\n')
								current_instance = instance_name.lower()
							v_write_content.append('tranif1 u_tran_'+instance_name.lower()+'_'+temp_table_row['pad_func'].lower()+'_'+i_opt)
							v_write_content.append(' (tb_'+instance_name.lower()+'_'+temp_table_row['pad_func'].lower()+', ')
							v_write_content.append(temp_table_row[i_opt]+', active_'+instance_name.lower()+'_'+i_opt+' );\n')
				for i_write_content in v_write_content:
					v_wire_file.write(i_write_content)
				v_wire_file.close()
				v_include_file.write('`include "'+v_wire_file_name+'"\n')
				#pinmux select
				#-------------------------------------------------------------------------
				module_name = temp_table.at[0, 'module'] # second row, second column
				c_write_content.append('static inline void sys_set_pinmux_'+module_name.lower()+' (uint32_t inst_name, uint32_t opt_index)'+'\n') 
				c_write_content.append('{\n') 
				opt_index = 0
				for i_opt in opt_list:
					#Check under this option, whether there is any valid function, if no then continue
					option_valid = 0
					for temp_table_row_index, temp_table_row in temp_table.iterrows() :
						if str(temp_table_row[i_opt]) != 'nan':
							option_valid = 1
							break
					if option_valid == 0:
						opt_index += 1
						continue
					#Check end
					c_write_content.append('    if(opt_index=='+str(opt_index)+') \n    {\n')
					current_instance = ''
					for temp_table_row_index, temp_table_row in temp_table.iterrows() :
						inst_name = temp_table_row['instance']
						#if current_instance == '' :
						#	current_instance = inst_name
						if current_instance != inst_name : # a new instance "if" block
							if current_instance != '' :
								c_write_content.append('            return;\n')
								c_write_content.append('        }\n')
							c_write_content.append('        if(inst_name==INST_'+str(inst_name)+')\n        {\n')
						if str(temp_table_row[i_opt]) != 'nan':
							if temp_table_row['domain'] == 'soc':
								#c_write_content.append('    if((inst_name==INST_'+str(inst_name)+')&(opt_index=='+str(opt_index)+')){\n')
								c_write_content.append('            REG32(MAP_IOC_'+temp_table_row[i_opt]+'_FUNC_CTL)=')
								c_write_content.append('REG32(MAP_IOC_'+temp_table_row[i_opt]+'_FUNC_CTL)')
								if temp_table_row['alt'] == 32 :
									c_write_content.append('|0x00000100; ') 
								else :
									c_write_content.append('&0xFFFFFFE0|'+str(temp_table_row['alt'])+'; ') 
								c_write_content.append('//'+temp_table_row['pad_func']+'\n') 
								if 'Y' in temp_table_row[i_opt]:
									c_write_content.append('            REG32(MAP_PMIC_IOC_'+temp_table_row[i_opt]+'_FUNC_CTL)=')
									c_write_content.append('REG32(MAP_PMIC_IOC_'+temp_table_row[i_opt]+'_FUNC_CTL)')
									c_write_content.append('&0xFFFFFFFC|3; ') 
									c_write_content.append('//MUX to SoC\n') 
								if 'Z' in temp_table_row[i_opt]:
									c_write_content.append('            REG32(MAP_BATT_IOC_'+temp_table_row[i_opt]+'_FUNC_CTL)=')
									c_write_content.append('REG32(MAP_BATT_IOC_'+temp_table_row[i_opt]+'_FUNC_CTL)')
									c_write_content.append('&0xFFFFFFFC|3; ') 
									c_write_content.append('//MUX to SoC\n') 
								#c_write_content.append('    }\n') 
							elif temp_table_row['domain'] == 'pmic':
								#c_write_content.append('    if((inst_name==INST_'+str(inst_name)+')&(opt_index=='+str(opt_index)+')){\n')
								c_write_content.append('            REG32(MAP_PMIC_IOC_'+temp_table_row[i_opt]+'_FUNC_CTL)=')
								c_write_content.append('REG32(MAP_PMIC_IOC_'+temp_table_row[i_opt]+'_FUNC_CTL)')
								c_write_content.append('&0xFFFFFFFC|'+str(temp_table_row['alt'])+'; ') 
								c_write_content.append('//'+temp_table_row['pad_func']+'\n') 
								#c_write_content.append('    }\n') 
							elif temp_table_row['domain'] == 'batt':
								#c_write_content.append('    if((inst_name==INST_'+str(inst_name)+')&(opt_index=='+str(opt_index)+')){\n')
								c_write_content.append('            REG32(MAP_BATT_IOC_'+temp_table_row[i_opt]+'_FUNC_CTL)=')
								c_write_content.append('REG32(MAP_BATT_IOC_'+temp_table_row[i_opt]+'_FUNC_CTL)')
								c_write_content.append('&0xFFFFFFFC|'+str(temp_table_row['alt'])+'; ') 
								c_write_content.append('//'+temp_table_row['pad_func']+'\n') 
								#c_write_content.append('    }\n') 
						current_instance = inst_name
					c_write_content.append('            return;\n')
					c_write_content.append('        }\n')
					c_write_content.append('    }\n')
					opt_index += 1
					#if (re.match(r'\d+', temp_table_row['instance'])):
					#	print ((re.search(r'\d+', temp_table_row['instance'])).group(1))
					#else:
					#	print ('AAA')
				c_write_content.append('    PRINTF(\"Set pinmux addr fail inst_name = %d opt_index = %d \",inst_name,opt_index);\n')
				c_write_content.append('    sys_exit(FAIL);\n')
				c_write_content.append('}\n\n') 
				#-------------------------------------------------------------------------
				# pinmux reg addr get
				c_write_content.append('static inline uint32_t sys_get_pinmux_addr_'+module_name.lower()+' (uint32_t inst_name, char *pad_func, uint32_t opt_index)'+'\n') 
				c_write_content.append('{\n') 
				opt_index = 0
				for i_opt in opt_list:
					#Check under this option, whether there is any valid function, if no then continue
					option_valid = 0
					for temp_table_row_index, temp_table_row in temp_table.iterrows() :
						if str(temp_table_row[i_opt]) != 'nan':
							option_valid = 1
							break
					if option_valid == 0:
						opt_index += 1
						continue
					#Check end
					c_write_content.append('    if(opt_index=='+str(opt_index)+') \n    {\n')
					current_instance = ''
					for temp_table_row_index, temp_table_row in temp_table.iterrows() :
						inst_name = temp_table_row['instance']
						pad_function = temp_table_row['pad_func']
						if current_instance != inst_name : # a new instance "if" block
							if current_instance != '' :
								c_write_content.append('        }\n')
							c_write_content.append('        if(inst_name==INST_'+str(inst_name)+')\n        {\n')
						c_write_content.append('            if(strcmp(pad_func,\"'+pad_function+'\")==0){')
						if str(temp_table_row[i_opt]) != 'nan':
							c_write_content.append('return MAP_IOC_'+temp_table_row[i_opt]+'_FUNC_CTL;}\n')
						else:
							c_write_content.append('PRINTF(\"get pinmux addr fail\");') 
							c_write_content.append('sys_exit(FAIL);}\n') 
						current_instance = inst_name
					c_write_content.append('        }\n')
					c_write_content.append('    }\n')
					opt_index += 1
				c_write_content.append('    PRINTF(\"get pinmux addr fail\");\n') 
				c_write_content.append('    sys_exit(FAIL);\n') 
				c_write_content.append('}\n\n') 
				#-------------------------------------------------------------------------
				# pad ctl reg addr get
				c_write_content.append('static inline uint32_t sys_get_padctl_addr_'+module_name.lower()+' (uint32_t inst_name, char *pad_func, uint32_t opt_index)'+'\n') 
				c_write_content.append('{\n') 
				opt_index = 0
				for i_opt in opt_list:
					#Check under this option, whether there is any valid function, if no then continue
					option_valid = 0
					for temp_table_row_index, temp_table_row in temp_table.iterrows() :
						if str(temp_table_row[i_opt]) != 'nan':
							option_valid = 1
							break
					if option_valid == 0:
						opt_index += 1
						continue
					#Check end
					c_write_content.append('    if(opt_index=='+str(opt_index)+') \n    {\n')
					current_instance = ''
					for temp_table_row_index, temp_table_row in temp_table.iterrows() :
						inst_name = temp_table_row['instance']
						pad_function = temp_table_row['pad_func']
						if current_instance != inst_name : # a new instance "if" block
							if current_instance != '' :
								c_write_content.append('        }\n')
							c_write_content.append('        if(inst_name==INST_'+str(inst_name)+')\n        {\n')
						c_write_content.append('            if(strcmp(pad_func,\"'+pad_function+'\")==0){')
						if str(temp_table_row[i_opt]) != 'nan':
							c_write_content.append('return MAP_IOC_'+temp_table_row[i_opt]+'_PAD_CTL;}\n')
						else:
							c_write_content.append('PRINTF(\"get pinmux addr fail\");') 
							c_write_content.append('sys_exit(FAIL);}\n') 
						current_instance = inst_name
					c_write_content.append('        }\n')
					c_write_content.append('    }\n')
					opt_index += 1
				c_write_content.append('    PRINTF(\"get pinmux addr fail\");\n') 
				c_write_content.append('    sys_exit(FAIL);\n') 
				c_write_content.append('}\n\n') 
				#-------------------------------------------------------------------------
				c_header_file = open(c_header_file_name,'w')
				for i_write_content in c_write_content:
					c_header_file.write(i_write_content)
				c_header_file.close()
				if((current_module.lower() != 'adc') & (current_module.lower() != 'cmp')):
					c_include_file.write('#include \"io_config/'+c_header_file_name_short+'\"\n')
				#print (c_write_content)
				temp_table.drop(temp_table.index,inplace=True)
				temp_table = temp_table.append(pinmux_table_row, ignore_index=True)
				current_module = pinmux_table_row['module']
		c_include_file.write('#endif //__SYS_IO_CONFIG__H__\n')
		c_include_file.close()
		v_include_file.close()

		







if __name__ == '__main__':
	a=gen_tb_io()
	a.arg_h()
	a.run()
